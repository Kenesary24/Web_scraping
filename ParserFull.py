from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse
from random import uniform
from tqdm import tqdm
import os
import re
import time

# ——————————————————————————————————————————————————————————————
#  Настройки Selenium
# ——————————————————————————————————————————————————————————————

def init_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def human_delay(min_sec=3, max_sec=7):
    time.sleep(uniform(min_sec, max_sec))


# ——————————————————————————————————————————————————————————————
#  Работа с Excel
# ——————————————————————————————————————————————————————————————

def load_existing_links(filename="autosave_results.xlsx"):
    if not os.path.exists(filename):
        return set()
    try:
        wb = load_workbook(filename)
        ws = wb.active
        existing_links = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing_links.add(clean_link(row[0]))
        print(f"♻️ Загружено {len(existing_links)} уже обработанных ссылок")
        return existing_links
    except Exception as e:
        print(f"⚠️ Ошибка загрузки файла {filename}: {e}")
        return set()

def save_to_excel(data, filename="autosave_results.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Ссылка", "Телефон"])

    existing_links = set(row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0])
    added = 0
    for row in data:
        url = row[0]
        if url not in existing_links:
            ws.append(row)
            existing_links.add(url)
            added += 1

    wb.save(filename)
    print(f"💾 Добавлено {added} новых строк в {filename}")


# ——————————————————————————————————————————————————————————————
#  Парсинг страницы объявления
# ——————————————————————————————————————————————————————————————

def parse_listing(driver, url):
    driver.get(url)
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'div.offer__advert-title'))
    )
    click_empty_space(driver)
    phone = extract_phone(driver)
    return [url, phone]


def get_total_pages(driver, url):
    driver.get(url)
    time.sleep(3)
    try:
        pagination = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a.paginator__btn'))
        )
        last_page = max([int(p.text) for p in pagination if p.text.isdigit()])
        print(f"🔢 Всего найдено страниц: {last_page}")
        return last_page
    except Exception:
        print("❗ Не удалось найти пагинацию. Используем 1 страницу.")
        return 1


def click_empty_space(driver):
    try:
        webdriver.ActionChains(driver).move_by_offset(10, 10).click().perform()
        time.sleep(1)
        webdriver.ActionChains(driver).move_by_offset(-10, -10).perform()
    except:
        pass

def clean_link(link):
    try:
        return urlparse(link.strip()).scheme + "://" + urlparse(link.strip()).netloc + urlparse(link.strip()).path
    except:
        return link.strip()
   

# ——————————————————————————————————————————————————————————————
#  Сбор ссылок с карточек
# ——————————————————————————————————————————————————————————————

def collect_links_from_page(driver):
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a.a-card__title'))
        )
    except:
        print("❗ Объявления не загрузились")
        return []

    cards = driver.find_elements(By.CSS_SELECTOR, "a.a-card__title")
    links = []
    for card in cards:
        href = card.get_attribute("href")
        if href and "/a/show/" in href:
            if href.startswith("/"):
                href = "https://krisha.kz" + href
            links.append(href)
    print(f"🔗 Найдено ссылок: {len(links)}")
    return links

# ——————————————————————————————————————————————————————————————
#  Сбор номера с карточек
# ——————————————————————————————————————————————————————————————

def extract_phone(driver):
    try:
        show_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "show-phones"))
        )
        show_btn.click()
        time.sleep(2)

        phone_el = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.offer__contacts-phones > p"))
        )
        print(f"📞 Извлечённый номер: {phone_el.text}")
        return re.sub(r"\D+", "", phone_el.text)

    except Exception as e:
        print(f"⚠️ Ошибка получения номера: {e}")
        return "Ошибка"


def process_link_with_filter(link, existing_links):
    if clean_link(link) in existing_links:
        return None
    driver = init_driver()
    try:
        data = parse_listing(driver, link)
        if data:
            print(f"💾 Готовим к сохранению: {data}")
        return data
    except Exception as e:
        print(f"❌ Ошибка при обработке {link}: {e}")
        return None
    finally:
        driver.quit()

# ——————————————————————————————————————————————————————————————
#  Главная функция
# ——————————————————————————————————————————————————————————————

def main():
    url = input("Вставь ссылку на список объявлений: ").strip()
    existing_links = load_existing_links("autosave_results.xlsx")
    all_data = []

    driver = init_driver()
    total_pages = get_total_pages(driver, url)
    driver.quit()

    page_limit = input(f"Сколько страниц обработать? (максимум {total_pages}): ").strip()
    try:
        page_limit = int(page_limit)
        if page_limit > total_pages:
            page_limit = total_pages
    except:
        page_limit = total_pages

    try:
        for page in range(1, page_limit + 1):
            print(f"\n📄 Обработка страницы {page} из {page_limit}")
            page_url = f"{url}?page={page}"

            driver = init_driver()
            driver.get(page_url)
            human_delay(2, 4)
            links = collect_links_from_page(driver)
            driver.quit()

            links_to_process = [link for link in links if clean_link(link) not in existing_links]

            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {
                    executor.submit(process_link_with_filter, link, existing_links): link
                    for link in links_to_process
                }
                for future in tqdm(as_completed(futures), total=len(futures), desc="Обработка объявлений"):
                    result = future.result()
                    if result:
                        save_to_excel([result])  # сохраняем сразу
                        all_data.append(result)
                        existing_links.add(clean_link(result[0]))

    except KeyboardInterrupt:
        print("\n🛑 Прерывание пользователем. Сохраняю уже обработанные данные...")

    finally:
        if all_data:
            save_to_excel(all_data)
            print(f"✅ Финальное сохранение завершено. Обработано: {len(all_data)} записей.")
        else:
            print("⚠️ Нет данных для сохранения.")

if __name__ == "__main__":
    main()
